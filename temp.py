# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""
import openai
import pandas as pd
import math
import numpy as np
import re
from dotenv import load_dotenv
from openai import OpenAI
import requests
import difflib
import json
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.drawing.image import Image
import xlwings as xw
import streamlit as st

def AIresponse(Email,x):     #This is for the hours worked
    url="https://api.groq.com/openai/v1/chat/completions"
    headers={
         "Authorization": "Bearer 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",
         "Content-Type": "application/json"
             }

    data = {
         "model": "meta-llama/llama-4-maverick-17b-128e-instruct",
         "messages": [
           {"role": "user", "content": "Extract JSON with occupation, matched occupation-NO NUMBERS, and hours only digit,travel km for each occupant, travel time for each occupant, from this text only.Use my headings. SO the data will be [the amount of matched occupations,5]. PLEASE indicate travel time and km next to the matched occupation and in seperate rows.!!!!! Please return the Json data [ and NOTHING else,no other comments, only correct data!!!\n\n.would you be able to read the following and format a table in JSON format with the occupation, matched occupation ,  and  with the amount of hours, travel km, travel time" f"Email:\n{Email}" "\n and also match the occupation to one occupation choose obvious one from list.Only return hours and matching occupation"f"list:\n{x}"}
           ]
           }

    response = requests.post(url, headers=headers, json=data)
    content=response.json()["choices"][0]["message"]["content"].strip()
    if content.startswith("```json") or content.startswith("```"):
        content = content.strip("```json").strip("Here is the extracted data in JSON format:").strip("```").strip()
    print(content)
    json_data=json.loads(content)
    df=pd.DataFrame.from_records(json_data)
    return df


def AIresponse1(Email,x):   # this is for the client details
    url="https://api.groq.com/openai/v1/chat/completions"
    headers={
         "Authorization": "Bearer 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",
         "Content-Type": "application/json"
             }

    data = {
         "model": "meta-llama/llama-4-maverick-17b-128e-instruct",
         "messages": [
           {"role": "user", "content": "Extract JSON with stating suitable subject for job, Business name,client's personal name client phone number and cleint email.Please reply ONLY with JSON data."f"Email:\n{Email}.\n please only send data back in JSON format start with [ !NO COMMENTS, just correct data! If no known information just state unknown!"}
           ]
           }

    response = requests.post(url, headers=headers, json=data)
    content=response.json()["choices"][0]["message"]["content"].strip()
    if content.startswith("```json") or content.startswith("```"):
        content = content.strip("```json").strip("Here is the extracted data in JSON format:").strip("Here are the extracted data in JSON format with the required headings:").strip("```").strip()
    print(content)
    json_data=json.loads(content)
    df=pd.DataFrame.from_records(json_data)
    return df

def AIresponse2(Email):        # This is for the job summary 
    url="https://api.groq.com/openai/v1/chat/completions"
    headers={
         "Authorization": "Bearer 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",
         "Content-Type": "application/json"
             }

    data = {
         "model": "meta-llama/llama-4-maverick-17b-128e-instruct",
         "messages": [
           {"role": "user", "content": "Please give me a one sentence job summary in proffessional words given this email. please only give me the job summary back "f"Email:\n{Email}"}
           ]
           }

    response = requests.post(url, headers=headers, json=data)
    content=response.json()["choices"][0]["message"]["content"].strip()
    print(content)
    return content

def AIresponse3(Email):        # This is for the job summary 
    url="https://api.groq.com/openai/v1/chat/completions"
    headers={
         "Authorization": "Bearer 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",
         "Content-Type": "application/json"
             }

    data = {
         "model": "meta-llama/llama-4-maverick-17b-128e-instruct",
         "messages": [
           {"role": "user", "content": "Please give me the product numbers and Quantity of the materials required in JSON format ONLY-no other word, just correct data PLEASE [( . Please only MATERIALS and not labour!. Please make that distinction!  Also use my headings Code and QTY!  If no materials the just return EMPTY JSON.  "f"Email:\n{Email}"}
           ]
           }

    response = requests.post(url, headers=headers, json=data)
    content=response.json()["choices"][0]["message"]["content"].strip()
    if content.startswith("```json") or content.startswith("```"):
        content = content.strip("```json").strip("Here is the extracted data in JSON format:").strip("Here are the extracted data in JSON format with the required headings:").strip("```").strip()
    json_data=json.loads(content)
    df=pd.DataFrame.from_records(json_data)
    print(df)
    return df



def writeBOQ(start,BOQ,x1,summary, materials):
    wb=load_workbook(r"C:\Users\henryl\Videos\(2024-11-21) Bill of Quantities - Rev 35.xlsm", keep_vba=True)   #Load workbook
    
    
    
    ws=wb['Start']# 
    wx=wb['Short Quote']
    wy=wb['BOQ']
    wz=wb['P&G\'s']
    #search for words and their data
    ws['J11']=start.iloc[0][0]
    ws['J12']=start.iloc[0][1]
    ws['J14']="South Africa"
    ws['J15']="2025"
    ws['J16']="0"
    ws['J17']=datetime.today().strftime('%d/%m/%Y')
    
    wx['B5']=start.iloc[0][2]
    wx['E9']=start.iloc[0][4]
    wx['B9']=start.iloc[0][3]
    if not materials.empty:
        wx['C26']="Material cost"
        
        
        wx['C28']="Labour cost"
        
        wx['C30']="P&G's"
        wx['C31']="Travel time and cost"
        wx['G26']='=GETPIVOTDATA("TOTAL PRICE\n(in Currency)",\'BOQ Pivot\'!$B$3,"Sections","Material cost")'
        wx['G28']='=GETPIVOTDATA("TOTAL PRICE\n(in Currency)",\'BOQ Pivot\'!$B$3,"Sections","Labour cost")'
        wx['G30']='=GETPIVOTDATA("TOTAL PRICE\n(in Currency)",\'BOQ Pivot\'!$B$3,"Sections","P&G\'s")'
        wx['A26']='1'
        wx['B26']='sum'
        wx['A28']='1'
        wx['B28']='sum'
        wx['A30']='1'
        wx['B30']='sum'
        wx['A14']=summary.strip("Here is a one sentence job summary in professional words:").strip("Here is a one-sentence job summary in professional words:")
        wx['C33']="Please note"
        wx['C34']="We have not allowed for weekend or overtime work"
        wx['C35']="We have not allowed medical, induction or safety file costs"
        
    if materials.empty:
     
      wx['C26']="Labour cost"
      
      wx['C28']="P&G's"
      wx['C29']="Travel time and cost"
      wx['G26']='=GETPIVOTDATA("TOTAL PRICE\n(in Currency)",\'BOQ Pivot\'!$B$3,"Sections","Labour cost")'
      wx['G28']='=GETPIVOTDATA("TOTAL PRICE\n(in Currency)",\'BOQ Pivot\'!$B$3,"Sections","P&G\'s")'
      wx['A26']='1'
      wx['B26']='sum'
      wx['A28']='1'
      wx['B28']='sum'
      wx['A14']=summary.strip("Here is a one sentence job summary in professional words:").strip("Here is a one-sentence job summary in professional words:")
      wx['C31']="Please note"
      wx['C32']="We have not allowed for weekend or overtime work"
      wx['C33']="We have not allowed medical, induction or safety file costs"  
        
 
    #My email, but it can be change accordingly
    wx['B51']="Henry Louw"
    wx['E51']="henryl@ppetech.co.za"
    wx['B52']="087 057 4001"
    wx['E52']="087 057 4001"
    wx['B53']="Henry Louw"
    img=Image(r"C:\Users\henryl\Pictures\signature.JPG")
    wx.add_image(img, "B54")
  
    
    #search for words and their data
    Occupation=[]
    x=difflib.get_close_matches('Matched occupation',BOQ)
    y=difflib.get_close_matches('hours',BOQ)
    z=difflib.get_close_matches('Travel Time',BOQ)
    w=difflib.get_close_matches('Travel km',BOQ)
    m1=difflib.get_close_matches('QTY',materials)
    m2=difflib.get_close_matches('Code',materials)
    wz['E78']=BOQ[w[0]].iloc[0]
    t=np.shape(BOQ)[0]
    
    if not materials.empty:
        for i in range((np.shape(BOQ)[0])):
           if (BOQ[x[0]].iloc[i])!="None" or (BOQ[x[0]].iloc[i])!="nan":
              Occupation=difflib.get_close_matches(BOQ[x[0]].iloc[i],x1)
              wy[f'I{28+i}']=Occupation[0]
              wy[f'J{28+i}']="Site installation"
              wy[f'K{28+i}']="Local"
              wy[f'L{28+i}']="hr"
              wy[f'N{28+i}']="PPE"
              wy[f'E{28+i}']="Labour cost"
              wy[f'M{28+i}']=BOQ[y[0]].iloc[i]
              
              for l in range(411-1):
                 if  wz[f'C{l+1}'].value==Occupation[0]  and l>=43 and l<=76:
                     wz[f'E{l+1}']=BOQ[z[0]].iloc[i]
              
              
              
        for k in range((np.shape(materials)[0])):
             wy[f'E{28+t+k}']="Material cost"
             wy[f'M{28+t+k}']=materials[m1[0]].iloc[k]
             wy[f'H{28+t+k}']=materials[m2[0]].iloc[k]
              
           
    
    
    
    
    
    
    if materials.empty:
        for i in range((np.shape(BOQ)[0])):
           if (BOQ[x[0]].iloc[i])!="None" or (BOQ[x[0]].iloc[i])!="nan":
              Occupation=difflib.get_close_matches(BOQ[x[0]].iloc[i],x1)
              wy[f'I{28+i}']=Occupation[0]
              wy[f'J{28+i}']="Site installation"
              wy[f'K{28+i}']="Local"
              wy[f'L{28+i}']="hr"
              wy[f'N{28+i}']="PPE"
              wy[f'E{28+i}']="Labour cost"
              wy[f'M{28+i}']=BOQ[y[0]].iloc[i]
          
          #Add travelling time
           for k in range(411-1):
              if  wz[f'C{k+1}'].value==Occupation[0]  and k>=43 and k<=76:
                  wz[f'E{k+1}']=BOQ[z[0]].iloc[i]
   
    
    
    
    
    wb.save(r"C:\Users\henryl\Videos\(2024-11-21) Bill of Quantities - Rev 35_updated.xlsm")
    wb.close() 
    
    wb1=xw.Book(r"C:\Users\henryl\Videos\(2024-11-21) Bill of Quantities - Rev 35_updated.xlsm")
    ws1=wb1.sheets['Short Quote']
    if materials.empty:
        ws1.range("16:22").api.Delete()
        ws1.range("28:43").api.Delete()
        #Clear all other formulas
        for i in range(50):
            if ws1.range(f'G{i+1}').value is None:
                ws1.range(f'H{i+1}').value=None
    
    if not materials.empty:
        ws1.range("16:22").api.Delete()
        ws1.range("30:43").api.Delete()
        #Clear all other formulas
        for i in range(50):
            if ws1.range(f'G{i+1}').value is None:
                ws1.range(f'H{i+1}').value=None
    
    
    wb1.save()
    wb1.close()
    
    
st.title("Generate your basic quote")  
st.text_area("Please place insert email or information")
if st.button:


    rates=pd.read_excel(r"C:\Users\henryl\Documents\book1.xlsx")
    
    x1=rates.iloc[:,0] 
    
    email=input()
    Start=AIresponse1(email,x1)
    print(Start.iloc[0][1])
    BOQ=AIresponse(email,x1)
    Summary_sentence=AIresponse2(email)
    Material=AIresponse3(email)
    
    
    x=difflib.get_close_matches('Matched occupation',BOQ)
    print(x)
    writeBOQ(Start,BOQ,x1,Summary_sentence, Material)
 




        
         